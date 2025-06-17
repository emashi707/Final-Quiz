from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('sagatave_eksamenam.xlsx')

# Select the relevant sheet
ws = wb['Lapa_0']

count = 0

# Loop through each row, starting from the second row to skip headers
for row in range(2, ws.max_row + 1):
    priority = ws['H' + str(row)].value  # Column H is "Prioritāte"
    delivery_date = ws['J' + str(row)].value  # Column J is "Piegādes datums"
    if priority == 'High' and delivery_date:
        # Check if delivery date is in the year 2015
        if delivery_date.year == 2015:
            count += 1

print("Number of matching records:", count)