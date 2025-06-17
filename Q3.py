from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('sagatave_eksamenam.xlsx')

# Select the relevant sheet
ws = wb['Lapa_0']

count = 0

# Loop through each row, starting from the second row (assuming first row is header)
for row in range(2, ws.max_row + 1):
    address = ws['D' + str(row)].value  # Column D is "Adrese"
    city = ws['E' + str(row)].value     # Column E is "Pilsēta"
    if address == "Adulienas iela" and city in ["Valmiera", "Saulkrasti"]:
        count += 1

print("Number of matching records:", count)