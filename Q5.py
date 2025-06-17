from openpyxl import load_workbook
import math

# Load the Excel file
wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

# Headers are in row 3
header_row = 3
headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]

# Find column indices
klients_col = headers.index('Klients') + 1
skaits_col = headers.index('Skaits') + 1
kopa_col = headers.index('Kopā') + 1

total_sum = 0

for row in range(header_row + 1, ws.max_row + 1):
    klients = ws.cell(row=row, column=klients_col).value
    skaits = ws.cell(row=row, column=skaits_col).value
    kopa = ws.cell(row=row, column=kopa_col).value
    try:
        skaits_num = float(skaits)
        kopa_num = float(kopa)
    except (TypeError, ValueError):
        continue
    if klients == 'Korporatīvais' and 40 <= skaits_num <= 50:
        total_sum += kopa_num

# Round down to integer
result = math.floor(total_sum)
print("Total sum (rounded down):", result)