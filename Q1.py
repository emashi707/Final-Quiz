from openpyxl import load_workbook


wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_1']

count = 0
for row in range(2, ws.max_row + 1):
    address = ws['A' + str(row)].value
    value = ws['B' + str(row)].value

    if address and isinstance(value, (int, float)):
        if address.startswith('Ain') and value < 40:
            count += 1

print("Number of matching records:", count)
