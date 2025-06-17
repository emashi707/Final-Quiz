from openpyxl import load_workbook
import math

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

# Headers are in row 3
header_row = 3
headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]

product_col = headers.index('Produkts') + 1
price_col = headers.index('Cena') + 1

total = 0
count = 0

for row in range(header_row + 1, ws.max_row + 1):
    product = ws.cell(row=row, column=product_col).value
    price = ws.cell(row=row, column=price_col).value
    if product and 'laserjet' in str(product).lower():
        try:
            price = float(price)
        except (TypeError, ValueError):
            continue
        total += price
        count += 1

if count > 0:
    average = total / count
    average_rounded_down = math.floor(average)
    print("Average Cena (rounded down):", average_rounded_down)
else:
    print("No matching records found.")